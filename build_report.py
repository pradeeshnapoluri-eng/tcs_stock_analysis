import pandas as pd
import numpy as np
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import warnings
warnings.filterwarnings("ignore")

wb = Workbook()

HEADER_FILL   = PatternFill("solid", fgColor="1565C0")
SUBHEAD_FILL  = PatternFill("solid", fgColor="1976D2")
ALT_FILL      = PatternFill("solid", fgColor="E3F2FD")
HEADER_FONT   = Font(bold=True, color="FFFFFF", size=11)
TITLE_FONT    = Font(bold=True, color="1565C0", size=13)
BORDER        = Border(
    left   = Side(style="thin"),
    right  = Side(style="thin"),
    top    = Side(style="thin"),
    bottom = Side(style="thin")
)

def style_sheet(ws, df, title):
    ws.append([title])
    ws.cell(1, 1).font = TITLE_FONT
    ws.append([])
    ws.append(df.columns.tolist())
    for col, val in enumerate(df.columns.tolist(), 1):
        cell = ws.cell(3, col)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center")
        cell.border = BORDER
    for row_idx, row in enumerate(df.values.tolist(), 4):
        fill = ALT_FILL if row_idx % 2 == 0 else PatternFill()
        for col_idx, val in enumerate(row, 1):
            cell = ws.cell(row_idx, col_idx, val)
            cell.fill = fill
            cell.border = BORDER
            cell.alignment = Alignment(horizontal="center")
    for col in ws.columns:
        max_len = max((len(str(c.value)) for c in col if c.value), default=10)
        ws.column_dimensions[get_column_letter(col[0].column)].width = min(max_len + 4, 30)

files = {
    "Overview"      : ("sql_overall_stats.csv",   "TCS Overall Statistics"),
    "Raw Data"      : ("TCS_cleaned.csv",          "TCS Cleaned Stock Data"),
    "Yearly"        : ("sql_yearly_stats.csv",     "Yearly Price Summary"),
    "Monthly"       : ("sql_monthly_stats.csv",    "Monthly Price Summary"),
    "Quarterly"     : ("sql_quarterly_stats.csv",  "Quarterly Performance"),
    "Top10 High"    : ("sql_top10_high.csv",       "Top 10 Highest Close Days"),
    "Top10 Volume"  : ("sql_top10_volume.csv",     "Top 10 Highest Volume Days"),
    "Dividends"     : ("sql_dividends.csv",        "Dividend Events"),
    "Best Months"   : ("sql_best_months.csv",      "Best Performing Months"),
    "Price Bands"   : ("sql_price_bands.csv",      "Price Band Distribution"),
    "Signals"       : ("sql_signals.csv",          "Buy vs Sell Signal Days"),
    "ML Results"    : ("tcs_model_results.csv",    "ML Model Performance"),
    "Feature Imp"   : ("tcs_feature_importance.csv","Feature Importance"),
    "Predictions"   : ("tcs_predictions.csv",      "Model Predictions vs Actual")
}

first = True
for sheet_name, (csv_file, title) in files.items():
    try:
        df = pd.read_csv(csv_file)
        if first:
            ws = wb.active
            ws.title = sheet_name
            first = False
        else:
            ws = wb.create_sheet(sheet_name)
        if sheet_name == "Raw Data":
            df = df.head(500)
        style_sheet(ws, df.round(4), title)
        print("Sheet done:", sheet_name)
    except Exception as e:
        print("Skipped:", sheet_name, "—", e)

wb.save("TCS_Stock_Report.xlsx")
print("build_report.py done — TCS_Stock_Report.xlsx saved")
